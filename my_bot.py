from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes, ConversationHandler
from openpyxl import Workbook, load_workbook
import os

TOKEN = "7603428852:AAHRS8vizKnQuFKJ3EShsMeLjgP3OM3PsYw"

ASK_NAME, ASK_AGE, ASK_DAY, ASK_TIME = range(4)
TOUR_NAME, TOUR_BIRTH, TOUR_PHONE, TOUR_CITY = range(4, 8)

ADMIN_ID = 2077512321
TRAINER_IDS = [2077512321]

days = ["Понедельник", "Вторник", "Среда", "Четверг"]
times = ["9:00-10:30", "10:30-12:00", "15:00-16:30", "16:30-18:00"]

def save_to_excel(filename, data_row, headers=None):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        if headers:
            ws.append(headers)
    else:
        wb = load_workbook(filename)
        ws = wb.active
    ws.append(data_row)
    wb.save(filename)

async def send_main_menu(message):
    keyboard = [
        [InlineKeyboardButton("👨‍💼 Руководство", callback_data="management")],
        [InlineKeyboardButton("🏋️‍♂️ Тренеры", callback_data="coaches")],
        [InlineKeyboardButton("🎮 Дисциплины", callback_data="disciplines")],
        [InlineKeyboardButton("✅ Записаться на занятие", callback_data="start_signup")],
        [InlineKeyboardButton("🏆 Регистрация на турниры", callback_data="tournaments")],
    ]
    await message.reply_text("👋 Привет! Добро пожаловать!\nВыбери раздел:", reply_markup=InlineKeyboardMarkup(keyboard))

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await send_main_menu(update.message)

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "management":
        keyboard = [
            [InlineKeyboardButton("Председатель", callback_data="chairman")],
            [InlineKeyboardButton("Исполнительный директор", callback_data="executive")],
            [InlineKeyboardButton("Директор турниров", callback_data="tournament_director")],
            [InlineKeyboardButton("Администратор турниров", callback_data="tournament_admin")],
            [InlineKeyboardButton("🏠 Главное меню", callback_data="back_main")]
        ]
        await query.message.reply_text("Выбери руководителя:", reply_markup=InlineKeyboardMarkup(keyboard))

    elif query.data == "chairman":
        keyboard = [[InlineKeyboardButton("🔙 Назад", callback_data="management")]]
        await query.message.reply_photo(
            photo="https://i.imgur.com/FYfd7YA.jpg",
            caption="👨‍💼 **Председатель** — Надырханов Дмитрий Сергеевич",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif query.data == "executive":
        keyboard = [[InlineKeyboardButton("🔙 Назад", callback_data="management")]]
        await query.message.reply_photo(
            photo="https://i.imgur.com/sBF4K9U.jpg",
            caption="👔 **Исполнительный директор** — Мелузов Константин Анатольевич",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif query.data == "tournament_director":
        keyboard = [[InlineKeyboardButton("🔙 Назад", callback_data="management")]]
        await query.message.reply_photo(
            photo="https://i.imgur.com/962FLlS.jpg",
            caption="🏆 **Директор турниров** — Матюшкин Радик Маратович",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif query.data == "tournament_admin":
        keyboard = [[InlineKeyboardButton("🔙 Назад", callback_data="management")]]
        await query.message.reply_photo(
            photo="https://i.imgur.com/LAeiXQx.jpg",
            caption="🗂 **Администратор турниров** — Мищенко Евгений Сергеевич",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif query.data == "coaches":
        keyboard = [
            [InlineKeyboardButton("Главный тренер", callback_data="head_coach")],
            [InlineKeyboardButton("Тренер по дронам", callback_data="drone_coach")],
            [InlineKeyboardButton("🏠 Главное меню", callback_data="back_main")]
        ]
        await query.message.reply_text("Выбери тренера:", reply_markup=InlineKeyboardMarkup(keyboard))

    elif query.data == "head_coach":
        keyboard = [[InlineKeyboardButton("🔙 Назад", callback_data="coaches")]]
        await query.message.reply_photo(
            photo="https://i.imgur.com/WmM5c87.jpg",
            caption="👨‍🏫 **Главный тренер** — Мелузов Константин\nСпециализация: CS 2.",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif query.data == "drone_coach":
        keyboard = [[InlineKeyboardButton("🔙 Назад", callback_data="coaches")]]
        await query.message.reply_photo(
            photo="https://i.imgur.com/LAeiXQx.jpg",
            caption="🚁 **Тренер по дронам** — Мищенко Евгений\nСпециализация: DCL.",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif query.data == "disciplines":
        keyboard = [
            [InlineKeyboardButton("Гонки дронов", callback_data="drones")],
            [InlineKeyboardButton("Боевая арена", callback_data="arena")],
            [InlineKeyboardButton("Головоломки", callback_data="puzzles")],
            [InlineKeyboardButton("Стратегия в реальном времени", callback_data="rts")],
            [InlineKeyboardButton("Технический симулятор", callback_data="tech_sim")],
            [InlineKeyboardButton("Спортивный симулятор", callback_data="sport_sim")],
            [InlineKeyboardButton("Файтинг", callback_data="fighting")],
            [InlineKeyboardButton("🏠 Главное меню", callback_data="back_main")]
        ]
        await query.message.reply_text("Выберите дисциплину:", reply_markup=InlineKeyboardMarkup(keyboard))

    elif query.data in ["drones", "arena", "puzzles", "rts", "tech_sim", "sport_sim", "fighting"]:
        descriptions = {
            "drones": "🚁 **Гонки дронов** — скоростная дисциплина, представляющая собой захватывающие соревнования FPV-квадрокоптеров на трассах.",
            "arena": "⚔ **Боевая арена** — командные сражения с целью разрушить главное здание соперника.",
            "puzzles": "🧩 **Соревновательные головоломки** — решение логических задач на скорость.",
            "rts": "🕹 **Стратегия в реальном времени** — управление армиями и ресурсами в динамичном темпе.",
            "tech_sim": "🔧 **Технический симулятор** — управление техникой по спортивным правилам.",
            "sport_sim": "⚽ **Спортивный симулятор** — воспроизведение спортивных игр по настоящим правилам.",
            "fighting": "🥊 **Файтинг** — рукопашные бои двух персонажей на арене.",
        }
        photos = {
            "drones": "https://i.imgur.com/K5JjZ0B.jpg",
            "arena": "https://i.imgur.com/QO1i48o.jpg",
            "puzzles": "https://i.imgur.com/0cybYkZ.jpg",
            "rts": "https://i.imgur.com/fosdq3a.jpg",
            "tech_sim": "https://i.imgur.com/yWk7lgY.jpg",
            "sport_sim": "https://i.imgur.com/v58jK3i.jpg",
            "fighting": "https://i.imgur.com/i0lTqTU.jpg",
        }
        keyboard = [
            [InlineKeyboardButton("🔙 Назад", callback_data="disciplines")],
            [InlineKeyboardButton("🏠 Главное меню", callback_data="back_main")]
        ]
        await query.message.reply_photo(
            photo=photos[query.data],
            caption=descriptions[query.data],
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif query.data == "tournaments":
        keyboard = [
            [InlineKeyboardButton("23 июля: Dota 2", callback_data="july_dota")],
            [InlineKeyboardButton("23 июля: Шашки", callback_data="july_checkers")],
            [InlineKeyboardButton("8 августа: FC 25", callback_data="aug_fc")],
            [InlineKeyboardButton("8 августа: Шашки", callback_data="aug_checkers")],
            [InlineKeyboardButton("🏠 Главное меню", callback_data="back_main")]
        ]
        await query.message.reply_text("Выберите турнир:", reply_markup=InlineKeyboardMarkup(keyboard))

    elif query.data in ["july_dota", "july_checkers", "aug_fc", "aug_checkers"]:
        await query.message.reply_text("Введите ваше ФИО для регистрации на турнир:")
        context.user_data["tournament"] = query.data
        return TOUR_NAME

    elif query.data == "start_signup":
        await query.message.reply_text("Введите ваше ФИО:")
        return ASK_NAME

    elif query.data == "back_main":
        await send_main_menu(query.message)

async def ask_age(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["name"] = update.message.text
    await update.message.reply_text("Укажите ваш возраст:")
    return ASK_AGE

async def ask_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["age"] = update.message.text
    keyboard = [[KeyboardButton(day)] for day in days]
    await update.message.reply_text("Выберите день недели:", reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True))
    return ASK_DAY

async def ask_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["day"] = update.message.text
    keyboard = [[KeyboardButton(time)] for time in times]
    await update.message.reply_text("Выберите время:", reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True))
    return ASK_TIME

async def confirm_signup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["time"] = update.message.text
    name = context.user_data["name"]
    age = context.user_data["age"]
    day = context.user_data["day"]
    time = context.user_data["time"]

    save_to_excel("training_signups.xlsx", [name, age, day, time], ["ФИО", "Возраст", "День", "Время"])
    for trainer_id in TRAINER_IDS:
        await context.bot.send_message(trainer_id, f"👤 Новая запись на занятие:\nФИО: {name}\nВозраст: {age}\nДень: {day}\nВремя: {time}")

    await update.message.reply_text("✅ Вы успешно записаны!", reply_markup=ReplyKeyboardMarkup([["/start"]], resize_keyboard=True))
    return ConversationHandler.END

async def tour_birth(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["name"] = update.message.text
    await update.message.reply_text("Укажите дату рождения:")
    return TOUR_BIRTH

async def tour_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["birth"] = update.message.text
    await update.message.reply_text("Укажите номер телефона:")
    return TOUR_PHONE

async def tour_city(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["phone"] = update.message.text
    await update.message.reply_text("Укажите город проживания:")
    return TOUR_CITY

async def tour_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["city"] = update.message.text
    name = context.user_data["name"]
    birth = context.user_data["birth"]
    phone = context.user_data["phone"]
    city = context.user_data["city"]
    tournament = context.user_data["tournament"]

    filename = f"{tournament}_signups.xlsx"
    save_to_excel(filename, [name, birth, phone, city], ["ФИО", "Дата рождения", "Телефон", "Город"])

    for trainer_id in TRAINER_IDS:
        await context.bot.send_message(trainer_id, f"👤 Новая регистрация на турнир ({tournament}):\nФИО: {name}\nДата рождения: {birth}\nТелефон: {phone}\nГород: {city}")

    await update.message.reply_text("✅ Спасибо! Вы зарегистрированы на турнир!", reply_markup=ReplyKeyboardMarkup([["/start"]], resize_keyboard=True))
    return ConversationHandler.END

def main():
    app = ApplicationBuilder().token(TOKEN).build()

    conv_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(button_handler)],
        states={
            ASK_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_age)],
            ASK_AGE: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_day)],
            ASK_DAY: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_time)],
            ASK_TIME: [MessageHandler(filters.TEXT & ~filters.COMMAND, confirm_signup)],

            TOUR_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, tour_birth)],
            TOUR_BIRTH: [MessageHandler(filters.TEXT & ~filters.COMMAND, tour_phone)],
            TOUR_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, tour_city)],
            TOUR_CITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, tour_confirm)],
        },
        fallbacks=[]
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(conv_handler)
    app.add_handler(CallbackQueryHandler(button_handler))
    print("Бот запущен!")
    app.run_polling()

if __name__ == "__main__":
    main()
